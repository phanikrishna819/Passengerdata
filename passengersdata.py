import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font

def data():  
    print("enter the data\n")
    
    print("Passenger Name, ""Ticket number, " "Flight Number, " "Flight Name, ""Boarding Point, ""Destination, ""Departure Time, ""Phone Number")
    list = []
    ##n= int(input("enter number of passengers:\n"))
    ##for row in range(0,n):
    for i in range(0,8):
        element = input("")
        list.append(element)
    print(list)   
    for column in range(0,1):
        member =  list
        sheet.append(member)
    print("---Data Stored---")
    choice = input("would you like to enter the data yes/no\n")
    if(choice=="yes"):
        return data()
    File = input("enter the file name to store the data")
    wb.save(File+".xlsx")

bold = Font(bold = True)
print("would you like to enter the data yes/no\n")
wb = openpyxl.Workbook()
sheet = wb.active
c1 = sheet.cell(row=1, column =1,)
sheet.cell(row = 1, column = 1).font = bold
c1.value = "Passenger Name"
c2 = sheet.cell(row =1, column = 2)
sheet.cell(row = 1, column = 2).font = bold
c2.value = "Ticket Number"
c3 = sheet.cell(row =1, column = 3)
sheet.cell(row = 1, column = 3).font = bold
c3.value = "Flight Number"
c4 = sheet.cell(row =1, column = 4)
sheet.cell(row = 1, column = 4).font = bold
c4.value = "Flight Name"
c5 = sheet.cell(row =1, column = 5)
sheet.cell(row = 1, column = 5).font = bold
c5.value = "Boarding point"
c6 = sheet.cell(row =1, column = 6)
sheet.cell(row = 1, column = 6).font = bold
c6.value = "Destination"
c7 = sheet.cell(row =1, column = 7)
sheet.cell(row = 1, column = 7).font = bold
c7.value = "departure Time"
c8 = sheet.cell(row =1, column = 8)
sheet.cell(row = 1, column = 8).font = bold
c8.value = "Phone Number"

choice = input("")
if(choice=="yes"):
    data()
elif(choice=="no"):
    exit(0)
